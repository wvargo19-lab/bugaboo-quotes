import json, base64, sys, os
sys.path.insert(0, os.path.dirname(__file__))

def fill_workbook(snap, quote_num, quote_date):
    from openpyxl import load_workbook
    import io
    
    template_path = os.path.join(os.path.dirname(__file__), 'template.xlsx')
    wb = load_workbook(template_path)
    
    S = snap
    ph = S.get('ph', {})
    fv = lambda v: (float(str(v).replace('$','').replace(',','')) if v not in ('','',None) else 0)

    TO  = '📊 Takeoff & Quote'
    DB  = '📋 Dashboard'
    CQ2 = '📄 Customer Quote'

    def sv(sheet, addr, val):
        ws = wb[sheet]
        ws[addr].value = val

    proj = (S.get('client','') + (' — ' + S.get('service','') if S.get('service') else '')).strip(' — ')
    client   = S.get('client','')
    site     = S.get('site','')
    operator = S.get('operator','')
    notes    = S.get('notes','') or S.get('projectDesc','')

    sv(DB,'C3',proj); sv(DB,'C4',site); sv(DB,'C5',client); sv(DB,'C8',operator)
    if notes: sv(DB,'C24',notes)

    sv(TO,'C2',client); sv(TO,'H2',proj); sv(TO,'C3',site); sv(TO,'H3',operator)
    sv(TO,'H4',quote_date or ''); sv(TO,'H5',quote_num or '')
    if notes: sv(TO,'C6',notes)

    sv(TO,'C8', fv(S.get('markup',30))/100)
    sv(TO,'E8', fv(S.get('costRate',35)))
    sv(TO,'G8', fv(ph.get('mob',{}).get('machRate',145)))
    sv(TO,'I8', fv(ph.get('mob',{}).get('fuelRate',45)))
    sv(TO,'K8', fv(ph.get('mob',{}).get('dumpRate',80)))
    sv(TO,'M8', fv(ph.get('mob',{}).get('delRate',120)))
    sv(TO,'O8', fv(S.get('targetMargin',35))/100)
    sv(TO,'P8', fv(S.get('billRate',85)))

    demo = ph.get('demo',{})
    if fv(demo.get('labHrs')): sv(TO,'H11',fv(demo.get('labHrs')))
    if fv(demo.get('dumpQty')):
        sv(TO,'E14',fv(demo.get('dumpQty'))); sv(TO,'F14',fv(demo.get('dumpRate',80)))

    grade = ph.get('grade',{})
    if fv(grade.get('labHrs')): sv(TO,'H18',fv(grade.get('labHrs')))

    irrig = ph.get('irrig',{})
    if fv(irrig.get('labHrs')): sv(TO,'H25',fv(irrig.get('labHrs')))
    if fv(irrig.get('matCost')):
        sv(TO,'C25','Irrigation materials'); sv(TO,'E25',1); sv(TO,'F25',fv(irrig.get('matCost')))

    hard = ph.get('hard',{})
    if fv(hard.get('labHrs')):  sv(TO,'H32',fv(hard.get('labHrs')))
    if fv(hard.get('machHrs')): sv(TO,'J32',fv(hard.get('machHrs')))
    for i,line in enumerate(hard.get('lines',[])):
        r=32+i
        if r>37 or not fv(line.get('qty')): continue
        sv(TO,f'C{r}',line.get('desc','') or f'Hardscape {i+1}')
        sv(TO,f'D{r}',line.get('unit','sqft'))
        sv(TO,f'E{r}',fv(line.get('qty'))); sv(TO,f'F{r}',fv(line.get('cu')))

    soft = ph.get('soft',{})
    if fv(soft.get('labHrs')): sv(TO,'H41',fv(soft.get('labHrs')))
    next_soft=41
    for i,line in enumerate(soft.get('lines',[])):
        r=41+i
        if r>45 or not fv(line.get('qty')): continue
        sv(TO,f'C{r}',line.get('desc','') or f'Softscape {i+1}')
        sv(TO,f'D{r}',line.get('unit','sqft'))
        sv(TO,f'E{r}',fv(line.get('qty'))); sv(TO,f'F{r}',fv(line.get('cu')))
        next_soft=r+1

    for mat in S.get('xMats',[]):
        if next_soft>45 or not fv(mat.get('qty',0)): continue
        sv(TO,f'C{next_soft}',mat.get('type','Material'))
        sv(TO,f'D{next_soft}',mat.get('unit','yards'))
        sv(TO,f'E{next_soft}',fv(mat.get('qty'))); sv(TO,f'F{next_soft}',fv(mat.get('cost')))
        next_soft+=1

    stru = ph.get('stru',{})
    if fv(stru.get('labHrs')): sv(TO,'H49',fv(stru.get('labHrs')))
    for i,line in enumerate(stru.get('lines',[])):
        r=49+i
        if r>52 or not fv(line.get('qty')): continue
        sv(TO,f'C{r}',line.get('desc','') or f'Structure {i+1}')
        sv(TO,f'D{r}',line.get('unit','each'))
        sv(TO,f'E{r}',fv(line.get('qty'))); sv(TO,f'F{r}',fv(line.get('cu')))

    mob = ph.get('mob',{})
    if fv(mob.get('machHrs')):   sv(TO,'J56',fv(mob.get('machHrs')))
    if fv(mob.get('fuelLoads')): sv(TO,'E57',fv(mob.get('fuelLoads'))); sv(TO,'F57',fv(mob.get('fuelRate',45)))
    if fv(mob.get('dumpQty')):   sv(TO,'E58',fv(mob.get('dumpQty')));   sv(TO,'F58',fv(mob.get('dumpRate',80)))
    if fv(mob.get('delQty')):    sv(TO,'E59',fv(mob.get('delQty')));    sv(TO,'F59',fv(mob.get('delRate',120)))
    if fv(mob.get('permits')):   sv(TO,'L60',fv(mob.get('permits')))
    if fv(mob.get('contHrs')):   sv(TO,'H61',fv(mob.get('contHrs')))

    sv(CQ2,'E5',quote_num or ''); sv(CQ2,'E6',quote_date or '')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def handler(event, context):
    try:
        body = json.loads(event.get('body','{}'))
        snap = body.get('snap', {})
        quote_num = body.get('quoteNum','')
        quote_date = body.get('date','')
        
        xlsx_bytes = fill_workbook(snap, quote_num, quote_date)
        b64 = base64.b64encode(xlsx_bytes).decode()
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({'file': b64})
        }
    except Exception as e:
        return {
            'statusCode': 500,
            'headers': {'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': str(e)})
        }

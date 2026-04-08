import json, base64, os, io, traceback, sys

def handler(event, context):
    # Handle OPTIONS preflight
    if event.get('httpMethod') == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type'
            },
            'body': ''
        }

    headers = {
        'Content-Type': 'application/json',
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'POST, OPTIONS',
        'Access-Control-Allow-Headers': 'Content-Type'
    }

    try:
        # Test openpyxl import first
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            return {
                'statusCode': 500,
                'headers': headers,
                'body': json.dumps({'error': 'openpyxl not installed: ' + str(e)})
            }

        # Parse body
        raw_body = event.get('body') or '{}'
        if event.get('isBase64Encoded'):
            raw_body = base64.b64decode(raw_body).decode('utf-8')
        
        body = json.loads(raw_body)
        snap = body.get('snap', {})
        quote_num = body.get('quoteNum', '')
        quote_date = body.get('date', '')

        # Load template
        template_path = os.path.join(os.path.dirname(__file__), 'template.xlsx')
        if not os.path.exists(template_path):
            return {
                'statusCode': 500,
                'headers': headers,
                'body': json.dumps({'error': f'Template not found at {template_path}'})
            }

        wb = load_workbook(template_path)

        S = snap
        ph = S.get('ph', {})

        def fv(v):
            try:
                if v in ('', None): return 0
                return float(str(v).replace('$','').replace(',',''))
            except:
                return 0

        TO  = '\U0001f4ca Takeoff & Quote'
        DB  = '\U0001f4cb Dashboard'
        CQ2 = '\U0001f4c4 Customer Quote'

        def sv(sheet, addr, val):
            if sheet in wb.sheetnames and val not in ('', None):
                wb[sheet][addr].value = val

        proj = (S.get('client','') + (' \u2014 ' + S.get('service','') if S.get('service') else '')).strip(' \u2014 ')
        client   = S.get('client','')
        site     = S.get('site','')
        operator = S.get('operator','')
        notes    = S.get('notes','') or S.get('projectDesc','')

        # Dashboard
        sv(DB,'C3',proj); sv(DB,'C4',site); sv(DB,'C5',client); sv(DB,'C8',operator)
        if notes: sv(DB,'C24',notes)

        # Takeoff header
        sv(TO,'C2',client); sv(TO,'H2',proj); sv(TO,'C3',site)
        sv(TO,'H3',operator); sv(TO,'H4',quote_date); sv(TO,'H5',quote_num)
        if notes: sv(TO,'C6',notes)

        # Rates
        sv(TO,'C8', fv(S.get('markup',30))/100)
        sv(TO,'E8', fv(S.get('costRate',35)))
        mob = ph.get('mob',{})
        sv(TO,'G8', fv(mob.get('machRate',145)))
        sv(TO,'I8', fv(mob.get('fuelRate',45)))
        sv(TO,'K8', fv(mob.get('dumpRate',80)))
        sv(TO,'M8', fv(mob.get('delRate',120)))
        sv(TO,'O8', fv(S.get('targetMargin',35))/100)
        sv(TO,'P8', fv(S.get('billRate',85)))

        # Demolition
        demo = ph.get('demo',{})
        if fv(demo.get('labHrs')): sv(TO,'H11',fv(demo.get('labHrs')))
        if fv(demo.get('dumpQty')):
            sv(TO,'E14',fv(demo.get('dumpQty')))
            sv(TO,'F14',fv(demo.get('dumpRate',80)))

        # Grading
        grade = ph.get('grade',{})
        if fv(grade.get('labHrs')): sv(TO,'H18',fv(grade.get('labHrs')))

        # Irrigation
        irrig = ph.get('irrig',{})
        if fv(irrig.get('labHrs')): sv(TO,'H25',fv(irrig.get('labHrs')))
        if fv(irrig.get('matCost')):
            sv(TO,'C25','Irrigation materials')
            sv(TO,'E25',1); sv(TO,'F25',fv(irrig.get('matCost')))

        # Hardscape
        hard = ph.get('hard',{})
        if fv(hard.get('labHrs')):  sv(TO,'H32',fv(hard.get('labHrs')))
        if fv(hard.get('machHrs')): sv(TO,'J32',fv(hard.get('machHrs')))
        for i,line in enumerate(hard.get('lines',[])):
            r = 32+i
            if r > 37 or not fv(line.get('qty')): continue
            sv(TO,f'C{r}', line.get('desc','') or f'Hardscape {i+1}')
            sv(TO,f'D{r}', line.get('unit','sqft'))
            sv(TO,f'E{r}', fv(line.get('qty')))
            sv(TO,f'F{r}', fv(line.get('cu')))

        # Softscape
        soft = ph.get('soft',{})
        if fv(soft.get('labHrs')): sv(TO,'H41',fv(soft.get('labHrs')))
        next_soft = 41
        for i,line in enumerate(soft.get('lines',[])):
            r = 41+i
            if r > 45 or not fv(line.get('qty')): continue
            sv(TO,f'C{r}', line.get('desc','') or f'Softscape {i+1}')
            sv(TO,f'D{r}', line.get('unit','sqft'))
            sv(TO,f'E{r}', fv(line.get('qty')))
            sv(TO,f'F{r}', fv(line.get('cu')))
            next_soft = r + 1

        for mat in S.get('xMats',[]):
            if next_soft > 45: break
            if not fv(mat.get('qty',0)): continue
            sv(TO,f'C{next_soft}', mat.get('type','Material'))
            sv(TO,f'D{next_soft}', mat.get('unit','yards'))
            sv(TO,f'E{next_soft}', fv(mat.get('qty')))
            sv(TO,f'F{next_soft}', fv(mat.get('cost')))
            next_soft += 1

        # Structures
        stru = ph.get('stru',{})
        if fv(stru.get('labHrs')): sv(TO,'H49',fv(stru.get('labHrs')))
        for i,line in enumerate(stru.get('lines',[])):
            r = 49+i
            if r > 52 or not fv(line.get('qty')): continue
            sv(TO,f'C{r}', line.get('desc','') or f'Structure {i+1}')
            sv(TO,f'D{r}', line.get('unit','each'))
            sv(TO,f'E{r}', fv(line.get('qty')))
            sv(TO,f'F{r}', fv(line.get('cu')))

        # Mobilization
        if fv(mob.get('machHrs')):   sv(TO,'J56',fv(mob.get('machHrs')))
        if fv(mob.get('fuelLoads')): sv(TO,'E57',fv(mob.get('fuelLoads'))); sv(TO,'F57',fv(mob.get('fuelRate',45)))
        if fv(mob.get('dumpQty')):   sv(TO,'E58',fv(mob.get('dumpQty')));   sv(TO,'F58',fv(mob.get('dumpRate',80)))
        if fv(mob.get('delQty')):    sv(TO,'E59',fv(mob.get('delQty')));    sv(TO,'F59',fv(mob.get('delRate',120)))
        if fv(mob.get('permits')):   sv(TO,'L60',fv(mob.get('permits')))
        if fv(mob.get('contHrs')):   sv(TO,'H61',fv(mob.get('contHrs')))

        # Customer Quote
        sv(CQ2,'E5', quote_num)
        sv(CQ2,'E6', quote_date)

        # Save to buffer
        buf = io.BytesIO()
        wb.save(buf)
        b64 = base64.b64encode(buf.getvalue()).decode()

        return {
            'statusCode': 200,
            'headers': headers,
            'body': json.dumps({'file': b64})
        }

    except Exception as e:
        tb = traceback.format_exc()
        print(f"EXPORT ERROR: {e}\n{tb}", file=sys.stderr)
        return {
            'statusCode': 500,
            'headers': headers,
            'body': json.dumps({'error': str(e), 'trace': tb[-500:]})
        }
